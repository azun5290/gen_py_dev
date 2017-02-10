# from xlrd import open_workbook
import openpyxl
import re
import collections

# prep/open workbook for reading
wb = openpyxl.load_workbook('tester_del_2_short.xlsx')
print(type(wb))
print(wb.get_sheet_names())

# prep wb for writing ---- or write to second sheet in same...

# separate possible wb obj manipulation way/lib
from xlrd import open_workbook
sheet = wb.get_active_sheet

sheet1 = wb.get_sheet_by_name(u'tester_del_2_short')
sheet1
print(type(sheet1))
print(sheet1.title)
print(sheet1['A1'].value)

# Prep dictionary
tempoDictio = {}
td = collections.defaultdict(list)

# some prints of stuff, values, etc.
sheet1.cell(row=1, column=2)
sheet1.cell(row=1, column=2).value

# a def tree sample
d = collections.defaultdict(list)
def tree():
    return collections.defaultdict(tree)
d['js'].append({'foo': 'bar'})
d['js'].append({'other': 'thing'})
print(d)
# end

# print some stuff from sheet1 but also append same values to td list initialized from collections earlier
for i in range(1, 12, 3):
        print(i, sheet1.cell(row=i, column=2).value)
        td['cell'].append({sheet1.cell(row=i, column=1).value : sheet1.cell(row=i, column=2).value})

        # tempoDictio.keys(sheet1.cell(row=i)).value, tempoDictio.values(sheet1.cell(column=2)).value
print(td).values()

# looping in range 'cause I don't know how to close loop for now --- needs fixing
for k_item in range(1, 4, 1):
# for k_item in (td)['cell']:
    # print single items in td named 'cell' just populated --- needs fixing with whole sequence of items later
    print('%'*88)
    sgl_lst_item = (td)['cell'].__getitem__(k_item)
    # print(td)['cell'].__getitem__(k_item)
    print(sgl_lst_item)

    # somewhere here we can start decomposing inner objects
    for td_elem1 in sgl_lst_item:
        print td_elem1

    # NOPE
    # for k, v in iter(sgl_lst_item.keys(), sgl_lst_item.values()):
      #  print(k, v)
    # for dict_1, dict_2 in (sgl_lst_item.value[0], sgl_lst_item.value[1]):
      #  print(dict_1, dict_2)
    print('%' * 88)

tup = tuple(sheet1['A1':'B10'])
for t in tup:
    print('='*64)
    print(t)

# for rowOfCellObjects in sheet1['A1':'B10']:
for rowOfCellObjects in tup:
    # playing with cell values
    for cellObj in rowOfCellObjects:

        print('--- START OF ROW OF VALUES ---' + '\n')
        print(cellObj.coordinate, cellObj.value)
        # for k, v in (cellObj.value[0], cellObj.value[1]):
          #  td[k].append(v)
        print('--- END OF ROW OF VALUES ---' + '\n')

print('FUFFFFFFFFFFFFAAAAAAAAAAAAAAAAAAAA')

# AGAIN II :: Next doesn't work because it uses other libraries to open object for reading
'''

keys = [sheet.cell(0, col_index).value for col_index in xrange(sheet.ncols)]

dict_list = []
for row_index in xrange(1, sheet.nrows):
    d = {keys[col_index]: sheet1.cell(row_index, col_index).value
         for col_index in xrange(sheet.ncols)}
    dict_list.append(d)
    print(d.keys(),d.values())
print dict_list

'''
print('Reading rows...')
for row in range(1, sheet1.max_row + 1, 1):
       # Each row in the spreadsheet has data for one adg
       adg_name = sheet1['A' + str(row)].value
       adg_info = sheet1['B' + str(row)].value

# tempoDictio.setdefault(adg_name, {})
# tempoDictio[adg_name].setdefault(adg_info, {'adg_name': 0})
# print(tempoDictio)

# OOORRRRR
# Next doesn't work because it is using other library to open wb as an object and change values !!!
'''
keys = [sheet1.cell(0, col_index).value for col_index in xrange(sheet1.ncols)]

dict_list = []
for row_index in xrange(1, sheet1.nrows):
    d = {keys[col_index]: sheet1.cell(row_index, col_index).value
         for col_index in xrange(sheet1.ncols)}
    dict_list.append(d)

print('NEW dictionary list should be here...')
print dict_list
'''

'''
NOT WORKING frnow...
sheet1 = wb.active
sheet1.columns[1]

for cellObj in sheet1.columns[1]:
    print(cellObj.value)

'''


'''
??
anotherSheet = wb.active
anotherSheet
'''