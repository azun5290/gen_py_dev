from openpyxl import load_workbook, Workbook

 
def read_xlsx(filename, read_intestation=True):
	"""
	Return a list of the elements in the xlsx file. Each element is stored in a
	list that contains the fields of the element.

	:param string filename: the name of the xlsx file to read.
	:param bool intestation: flag to indicate if the file contains the
		intestation.
	:rtype: list(list)
	"""

	data = []
	wb = load_workbook(filename)
	main_sheet_name = wb.get_sheet_names()[0]
	main_sheet = wb.get_sheet_by_name(main_sheet_name)


	for index_row, row in enumerate(main_sheet.iter_rows()):

		row_list = [cell.value for cell in row]

		if read_intestation and index_row == 0:
			intestation = row_list
		else:	
			data.append(row_list)

	return data, intestation

def save_xlsx(filename, data, intestation=None):
	"""
	Store the list of the elements to the xlsx file. Each element contains a list
	with the fields of the element.

	:param string filename: the name of the xlsx file where store the data.
	:param list(list) data: the data to store.
	:param list(string) intestation: the list of the string to write as
		intestation.
	"""
	
	wb = Workbook()
	main_sheet = wb.active
	main_sheet.title = filename

	starting_row = 1
	if intestation:
		for index_col, title in enumerate(intestation, start=1):
			main_sheet.cell(row=starting_row, column=index_col).value = title
		starting_row += 1

	for index_row, element in enumerate(data, start=starting_row):
		for index_col, field in enumerate(element, start=1):
			main_sheet.cell(row=index_row, column=index_col).value = field

	wb.save(filename)
